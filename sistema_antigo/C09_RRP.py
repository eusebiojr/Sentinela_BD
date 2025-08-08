from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from datetime import datetime, timedelta, time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from io import BytesIO
from pathlib import Path
import os
import time
from dotenv import load_dotenv
from office365.sharepoint.client_context import ClientContext
from office365.runtime.auth.user_credential import UserCredential



# ==============================================================
# Configuração Inicial
# ==============================================================

load_dotenv()  # carrega variáveis do .env
SITE_URL    = "https://suzano.sharepoint.com/sites/Controleoperacional"
SP_USER     = os.getenv("SP_USER")
SP_PASSWORD = os.getenv("SP_PASSWORD")
RAIZ_SP     = "/sites/Controleoperacional/Documentos Compartilhados/Bases de Dados"
BASE_SP     = "CREARE/RRP/C09"
BASE_REPORTS = r"C:\Users\tallespaiva\Suzano S A\Controle operacional - Bases de Dados\CREARE\RRP\C09\2025"
caminho_reports = r"C:\Users\tallespaiva\Suzano S A\Controle operacional - Bases de Dados\CREARE\RRP\C09\2025\Reports\base de dados reports.xlsx"
site_url = "https://suzano.sharepoint.com/sites/Controleoperacional"
list_name = "Desvios"
username = "eusebioagj@suzano.com.br"
password = "290422@Cc"

# Caminho para o ChromeDriver

CHROME_DRIVER_PATH = os.getenv("CHROME_DRIVER_PATH", "").strip()
if not CHROME_DRIVER_PATH:
    raise ValueError("Defina CHROME_DRIVER_PATH no .env apontando para chromedriver.exe")

p = Path(CHROME_DRIVER_PATH)
if not p.is_absolute():
    p = Path(__file__).parent.joinpath(p).resolve()
if not p.is_file():
    raise FileNotFoundError(f"ChromeDriver não encontrado em '{p}'")
CHROME_DRIVER_PATH = str(p)

DOWNLOAD_TIMEOUT = 300  # segundos

# ===================================================================
# Função de download do relatório C09
# ===================================================================

def baixar_relatorio_c09(data_inicial: datetime, data_final: datetime) -> str:
    """
    - Faz login no Frotalog (ID 'userName', NAME 'password', XPATH btn_ok.gif).
    - Navega em menus (Relatórios → Dirigibilidade → C09).
    - Seleciona empresa "RB - TRANSP. CELULOSE" no dropdown 'branchId'.
    - Preenche filtros de data (mesmos xpaths do antigo).
    - Seleciona formato XLSX (ID 'typeXLSX').
    - Clica em 'Visualizar Relatório', fecha a janela de relatório.
    - Clica em 'Listar Relatórios' e monitora até o status 'Pronto' e clica no link correto.
    - Baixa report.xlsx para ~/Downloads.
    Retorna o caminho absoluto do arquivo baixado.
    """
    options = webdriver.ChromeOptions()
    # Para rodar headless, descomente a linha abaixo:
    options.add_argument("--headless=new")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920x1080")
    options.add_argument("--no-sandbox")

    # Forçar download automático para ~/Downloads
    pasta_download = Path.home() / "Downloads"
    prefs = {
        "download.prompt_for_download": False,
        "download.default_directory": str(pasta_download),
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True
    }
    options.add_experimental_option("prefs", prefs)

    service = Service(CHROME_DRIVER_PATH)
    driver = webdriver.Chrome(service=service, options=options)
    wait = WebDriverWait(driver, 20)

    try:
        # -------------------------------------------------
        # 1) Abre o site e faz login
        # -------------------------------------------------
        driver.get("https://frotalog.com.br/")
        time.sleep(2)  # aguarda carregamento inicial

        driver.find_element(By.ID, "userName").send_keys(os.getenv("FROTA_USER"))
        driver.find_element(By.NAME, "password").send_keys(os.getenv("FROTA_PASSWORD"))
        driver.find_element(By.XPATH, "//input[contains(@src, 'btn_ok.gif')]").click()
        print("Login enviado. Verificando pop-up...")

        # Fecha possíveis alertas de JavaScript
        try:
            WebDriverWait(driver, 3).until(EC.alert_is_present())
            driver.switch_to.alert.accept()
            print("Alerta fechado.")
        except TimeoutException:
            pass

        # Fecha pop-up de sessão
        try:
            wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//input[@value='Continuar com a sessão atual']")
            )).click()
            print("Pop-up de sessão tratado.")
        except TimeoutException:
            print("Nenhum pop-up detectado. Seguindo normalmente.")

        # -------------------------------------------------
        # 2) Navega até o relatório C09
        # -------------------------------------------------
        wait.until(EC.frame_to_be_available_and_switch_to_it((By.NAME, "main")))
        print("Frame 'main' carregado com sucesso.")

        # HOVER em "Relatórios" e clica em "Dirigibilidade"
        menu = wait.until(EC.presence_of_element_located((By.XPATH, "//*[@id='menu']/a")))
        submenu = driver.find_element(By.XPATH, "/html/body/form/div[1]/div[1]/table/tbody/tr/td[5]/ul/li[1]/a")
        ActionChains(driver).move_to_element(menu).pause(1).move_to_element(submenu).click().perform()
        print("Entrou em Dirigibilidade.")
        time.sleep(2)

        # Clica em C09
        driver.find_element(By.XPATH, "//a[contains(text(), 'C09')]").click()
        print("Selecionou relatório C09.")
        time.sleep(3)

        # -------------------------------------------------
        # 3) Reentra no frame principal para preencher filtros
        # -------------------------------------------------
        driver.switch_to.default_content()
        wait.until(EC.frame_to_be_available_and_switch_to_it((By.NAME, "main")))
        print("Reentrou no frame 'main' após carregar o relatório.")

        # Espera o dropdown de empresa 'branchId'
        wait.until(EC.presence_of_element_located((By.ID, "branchId")))
        # Seleciona empresa "RB - TRANSP. CELULOSE"
        empresa = driver.find_element(By.ID, "branchId")
        for option in empresa.find_elements(By.TAG_NAME, "option"):
            if "RB - TRANSP. CELULOSE" in option.text:
                option.click()
                break
        print("Empresa selecionada.")
        time.sleep(1)

        # Seleciona modo 'Período' (ID 'range')
        wait.until(EC.element_to_be_clickable((By.ID, "range"))).click()
        print("Selecionado modo 'Período'.")

        # -------------------------------------------------
        # 4) Preenche datas de início e fim (mesmo código antigo)
        # -------------------------------------------------
        dia_ini = str(data_inicial.day)
        mes_ini = str(data_inicial.month)
        ano_ini = str(data_inicial.year)
        dia_fim = str(data_final.day)
        mes_fim = str(data_final.month)
        ano_fim = str(data_final.year)

        # Data inicial
        driver.find_element(By.XPATH, "//*[@id='periodo']/tbody/tr[2]/td[1]/input[1]").clear()
        driver.find_element(By.XPATH, "//*[@id='periodo']/tbody/tr[2]/td[1]/input[1]").send_keys(dia_ini)
        time.sleep(0.5)
        driver.find_element(By.XPATH, "//*[@id='periodo']/tbody/tr[2]/td[1]/input[2]").clear()
        driver.find_element(By.XPATH, "//*[@id='periodo']/tbody/tr[2]/td[1]/input[2]").send_keys(mes_ini)
        time.sleep(0.5)
        driver.find_element(By.XPATH, "//*[@id='periodo']/tbody/tr[2]/td[1]/input[3]").clear()
        driver.find_element(By.XPATH, "//*[@id='periodo']/tbody/tr[2]/td[1]/input[3]").send_keys(ano_ini)
        time.sleep(0.5)

        # Data final
        driver.find_element(By.XPATH, "//*[@id='periodo']/tbody/tr[2]/td[2]/input[1]").clear()
        driver.find_element(By.XPATH, "//*[@id='periodo']/tbody/tr[2]/td[2]/input[1]").send_keys(dia_fim)
        time.sleep(0.5)
        driver.find_element(By.XPATH, "//*[@id='periodo']/tbody/tr[2]/td[2]/input[2]").clear()
        driver.find_element(By.XPATH, "//*[@id='periodo']/tbody/tr[2]/td[2]/input[2]").send_keys(mes_fim)
        time.sleep(0.5)
        driver.find_element(By.XPATH, "//*[@id='periodo']/tbody/tr[2]/td[2]/input[3]").clear()
        driver.find_element(By.XPATH, "//*[@id='periodo']/tbody/tr[2]/td[2]/input[3]").send_keys(ano_fim)
        time.sleep(0.5)

        # -------------------------------------------------
        # 5) Seleciona formato XLSX e clica em "Visualizar Relatório"
        # -------------------------------------------------
        driver.find_element(By.ID, "typeXLSX").click()
        print("Selecionado formato XLSX.")

        driver.find_element(By.XPATH,
            "/html/body/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr/td[1]/table/"
            "tbody/tr[3]/td[2]/table/tbody/tr/td[2]/table/tbody/tr/td/form/table/tbody/"
            "tr[11]/td/table/tbody/tr[1]/td[1]/input"
        ).click()
        print("Botão 'Visualizar Relatório' clicado com sucesso.")
        time.sleep(2)

        # Aguarda nova janela abrir e fecha
        WebDriverWait(driver, 15).until(lambda d: len(d.window_handles) > 1)
        original = driver.window_handles[0]
        extra = driver.window_handles[1]
        driver.switch_to.window(extra)
        driver.close()
        driver.switch_to.window(original)
        print("Nova janela fechada. Retornando para página principal.")

        # -------------------------------------------------
        # 6) Reentra no frame e clica em "Listar Relatórios"
        # -------------------------------------------------
        driver.switch_to.default_content()
        wait.until(EC.frame_to_be_available_and_switch_to_it((By.NAME, "main")))

        driver.find_element(By.XPATH,
            "/html/body/table/tbody/tr[1]/td/table/tbody/tr/td[3]/table/tbody/tr[2]/"
            "td/table/tbody/tr/td[1]/table/tbody/tr/td[8]/a"
        ).click()
        print("Clicou em 'Listar Relatórios' para acompanhar o status.")

        # -------------------------------------------------
        # 7) Monitora status até ficar 'Pronto' e clica no link
        # -------------------------------------------------
        max_tentativas = 60
        intervalo_segundos = 10
        for tentativa in range(1, max_tentativas + 1):
            try:
                status = driver.find_element(
                    By.XPATH,
                    "/html/body/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr/"
                    "td[1]/table/tbody/tr[3]/td[2]/form/table/tbody/tr[2]/"
                    "td/table/tbody/tr[3]/td[8]"
                ).text
                if "Pronto" in status:
                    print(f"Relatório pronto na tentativa {tentativa}.")
                    link = driver.find_element(
                        By.XPATH,
                        "/html/body/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr/"
                        "td[1]/table/tbody/tr[3]/td[2]/form/table/tbody/tr[2]/"
                        "td/table/tbody/tr[3]/td[1]/a"
                    )
                    link.click()
                    print("Download iniciado. Aguardando conclusão…")
                    break
                else:
                    print(f"Tentativa {tentativa}: status atual é '{status}'…")
            except NoSuchElementException:
                print(f"Tentativa {tentativa}: elemento status não encontrado.")
            time.sleep(intervalo_segundos)
        else:
            raise TimeoutException("Relatório C09 não ficou pronto dentro do tempo limite.")

        # -------------------------------------------------
        # 8) Aguarda o arquivo 'report.xlsx' aparecer em ~/Downloads
        # -------------------------------------------------
        caminho_report = pasta_download / "report.xlsx"
        caminho_temp   = pasta_download / "report.xlsx.crdownload"
        segundos = 0
        while ((not caminho_report.exists()) or caminho_temp.exists()) and segundos < DOWNLOAD_TIMEOUT:
            time.sleep(1)
            segundos += 1

        if not caminho_report.exists():
            raise FileNotFoundError("Download não foi concluído dentro do tempo limite.")
        else:
            print("Download finalizado em:", caminho_report)
            return str(caminho_report)

    except Exception as e:
        print("Erro na etapa de download:", e)
        raise

    finally:
        driver.quit()

# ===================================================================
# Função de tratamento da planilha
# ===================================================================

def tratar_planilha_c09(caminho_arquivo_origem: str) -> BytesIO:
    """
    Lê 'report.xlsx', filtra pontos de interesse, agrupa entradas idênticas,
    calcula Tempo de Permanência, Trajetos (Carregado/Vazio) e Observações de SLA,
    e retorna um BytesIO com o Excel final formatado em tabela.
    """
    df = pd.read_excel(caminho_arquivo_origem)
    df["Ponto de Interesse"] = df["Ponto de Interesse"].astype(str).str.strip()

    pontos_desejados = [
        "PA AGUA CLARA",
        "Carregamento RRp",
        "Descarga INO", #Deletar no mes 07
        "Descarga Inocencia", #Deletar no mes 07
        "Manutenção Campo Grande",
        "Manutencao fabrica",
        "Manutenção Geral JSL RRP",
        "Oficina JSL",
        "Buffer frotas",
        "Abastecimento Frotas RRP",
        "Carregamento Fabrica RRP",
        "Posto Mutum",
        "Enlonamento RRP",
        "Terminal Inocencia",
        "Patio Carregado INO",
        "Patio Vazio INO",
        "Desc. INO"
    ]

    df = df[df["Ponto de Interesse"].isin(pontos_desejados)].copy()
    df = df.sort_values(by=["Veículo", "Data Entrada"])
    df = df[["Veículo", "Ponto de Interesse", "Data Entrada", "Data Saída", "Observações"]]

    df["Data Entrada"] = pd.to_datetime(df["Data Entrada"], dayfirst=True, errors="coerce")
    df["Data Saída"]   = pd.to_datetime(df["Data Saída"],   dayfirst=True, errors="coerce")

    def classificar_grupo(ponto: str) -> str:
        if ponto == "Carregamento RRp":
            return "Carregamento"
        #elif ponto == "Desc. INO": voltar dia 12/06
        elif ponto == "Descarga INO":
            return "Descarregamento"
        elif ponto in ["Manutenção Geral JSL RRP", "Manutenção Campo Grande", "Manutencao fabrica", "Oficina JSL"]:
            return "Manutenção"
        elif ponto in ["PA AGUA CLARA", "Buffer frotas", "Abastecimento Frotas RRP", "Posto Mutum", "Enlonamento RRP", "Patio Carregado INO", "Patio Vazio INO"]:
            return "Parada Operacional"
        elif ponto == "Carregamento Fabrica RRP":
            return "Fabrica"
        #elif ponto == "Terminal Inocencia": voltar dia 12/06
        elif ponto == "Descarga Inocencia":
            return "Terminal"
        else:
            return "Outros"

    df["Grupo"] = df["Ponto de Interesse"].apply(classificar_grupo)

    agrupados = []
    df_reset = df.reset_index(drop=True)
    idx = 0
    while idx < len(df_reset):
        atual = df_reset.iloc[idx]
        veic  = atual["Veículo"]
        ponto = atual["Ponto de Interesse"]
        grupo = atual["Grupo"]
        entrada = atual["Data Entrada"]
        saida   = atual["Data Saída"]

        j = idx + 1
        while j < len(df_reset) and df_reset.iloc[j]["Veículo"] == veic and df_reset.iloc[j]["Ponto de Interesse"] == ponto:
            saida = df_reset.iloc[j]["Data Saída"]
            j += 1

        agrupados.append({
            "Veículo": veic,
            "Ponto de Interesse": ponto,
            "Data Entrada": entrada,
            "Data Saída": saida,
            "Grupo": grupo,
            "Observações": atual.get("Observações", "")
        })
        idx = j

    df_ag = pd.DataFrame(agrupados)
    df_ag["Tempo Permanencia"] = (
        (df_ag["Data Saída"] - df_ag["Data Entrada"])
        .dt.total_seconds() / 3600
    )

    df_ag["Trajeto Carregado"] = pd.NA
    df_ag["Trajeto Vazio"]     = pd.NA
    df_ag["Observação"]        = ""

    sla_carga_h    = 1.0    * 1.3
    sla_tr_car_h   = 6.3667 * 1.3
    sla_desc_h     = 1.1833 * 1.3
    sla_tr_vazio_h = 6.0833 * 1.3

    def soma_justificativas(df_int: pd.DataFrame, inicio: int, fim: int):
        sub = df_int.iloc[inicio + 1 : fim]
        horas_mant = sub[sub["Grupo"] == "Manutenção"]["Tempo Permanencia"].sum()
        horas_oper = sub[sub["Grupo"] == "Parada Operacional"]["Tempo Permanencia"].sum()
        return horas_mant, horas_oper

    for i in range(len(df_ag)):
        atual = df_ag.iloc[i]
        veic  = atual["Veículo"]
        grupo = atual["Grupo"]
        saida_atual = atual["Data Saída"]

        if grupo == "Carregamento":
            if atual["Tempo Permanencia"] > sla_carga_h:
                df_ag.at[i, "Observação"] += (
                    f"Carga estourou SLA ({atual['Tempo Permanencia']:.2f}h > {sla_carga_h:.2f}h). "
                )
        elif grupo == "Descarregamento":
            if atual["Tempo Permanencia"] > sla_desc_h:
                df_ag.at[i, "Observação"] += (
                    f"Descarga estourou SLA ({atual['Tempo Permanencia']:.2f}h > {sla_desc_h:.2f}h). "
                )

        if grupo == "Carregamento":
            for j in range(i + 1, len(df_ag)):
                prox = df_ag.iloc[j]
                if prox["Veículo"] != veic:
                    break
                if prox["Grupo"] == "Descarregamento":
                    if prox["Data Entrada"] >= saida_atual:
                        delta_h = (prox["Data Entrada"] - saida_atual).total_seconds() / 3600
                        df_ag.at[i, "Trajeto Carregado"] = round(delta_h, 5)
                        if delta_h > sla_tr_car_h:
                            horas_mant, horas_oper = soma_justificativas(df_ag, i, j)
                            if (horas_mant + horas_oper) > 0:
                                df_ag.at[i, "Observação"] += (
                                    f"Trajeto Carregado longo ({delta_h:.2f}h > {sla_tr_car_h:.2f}h): "
                                    f"{horas_mant:.2f}h em Manutenção, {horas_oper:.2f}h em Parada Operacional. "
                                )
                            else:
                                df_ag.at[i, "Observação"] += (
                                    f"Trajeto Carregado longo ({delta_h:.2f}h > {sla_tr_car_h:.2f}h), sem justificativa. "
                                )
                    break

        elif grupo == "Descarregamento":
            for j in range(i + 1, len(df_ag)):
                prox = df_ag.iloc[j]
                if prox["Veículo"] != veic:
                    break
                if prox["Grupo"] == "Carregamento":
                    if prox["Data Entrada"] >= saida_atual:
                        delta_h = (prox["Data Entrada"] - saida_atual).total_seconds() / 3600
                        df_ag.at[i, "Trajeto Vazio"] = round(delta_h, 5)
                        if delta_h > sla_tr_vazio_h:
                            horas_mant, horas_oper = soma_justificativas(df_ag, i, j)
                            if (horas_mant + horas_oper) > 0:
                                df_ag.at[i, "Observação"] += (
                                    f"Trajeto Vazio longo ({delta_h:.2f}h > {sla_tr_vazio_h:.2f}h): "
                                    f"{horas_mant:.2f}h em Manutenção, {horas_oper:.2f}h em Parada Operacional. "
                                )
                            else:
                                df_ag.at[i, "Observação"] += (
                                    f"Trajeto Vazio longo ({delta_h:.2f}h > {sla_tr_vazio_h:.2f}h), sem justificativa. "
                                )
                    break

    df_ag["Observação"] = df_ag["Observação"].str.strip()
    df_ag["Veículo"] = df_ag["Veículo"].astype(str).apply(lambda x: x.split("-")[-1].strip())
    df_ag["Data Entrada"] = df_ag["Data Entrada"].dt.strftime("%d/%m/%Y %H:%M:%S")
    df_ag["Data Saída"]   = df_ag["Data Saída"].dt.strftime("%d/%m/%Y %H:%M:%S")

    buffer_excel = BytesIO()
    with pd.ExcelWriter(buffer_excel, engine="openpyxl") as writer:
        df_ag.to_excel(writer, index=False, sheet_name="Relatório")
        # não chamar writer.save()—o contexto fecha e salva

    buffer_excel.seek(0)
    wb = load_workbook(buffer_excel)
    ws = wb.active
    num_linhas = ws.max_row
    num_colunas = ws.max_column

    def coluna_letra(n):
        letra = ""
        while n > 0:
            n, resto = divmod(n - 1, 26)
            letra = chr(65 + resto) + letra
        return letra

    ultima_coluna = coluna_letra(num_colunas)
    ref = f"A1:{ultima_coluna}{num_linhas}"
    tabela = Table(displayName="TabelaRelatorio", ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tabela.tableStyleInfo = style
    ws.add_table(tabela)

    final_buffer = BytesIO()
    wb.save(final_buffer)
    final_buffer.seek(0)

    return final_buffer

def carregar_planilha(caminho_arquivo):
    try:
        df = pd.read_excel(caminho_arquivo, engine='openpyxl')
        df['Data Entrada'] = pd.to_datetime(df['Data Entrada'], format='%d/%m/%Y %H:%M:%S', dayfirst=True, errors='coerce')
        df['Data Saída'] = pd.to_datetime(df['Data Saída'], format='%d/%m/%Y %H:%M:%S', dayfirst=True, errors='coerce')
        print(f"Planilha '{os.path.basename(caminho_arquivo)}' carregada com sucesso.")
        print(df[['Data Entrada', 'Data Saída']].dtypes)
        print(df[['Data Entrada', 'Data Saída']].head())
        return df
    except FileNotFoundError:
        print(f"Arquivo '{caminho_arquivo}' não encontrado.")
    except Exception as e:
        print(f"Ocorreu um erro ao tentar abrir a planilha: {e}")

def TPV(df, poi, data):
    df['Data Entrada'] = pd.to_datetime(df['Data Entrada'], format='%d/%m/%Y %H:%M:%S', errors='coerce')
    df_filtrado = df[(df['Data Entrada'].dt.date == data) & (df['Ponto de Interesse'] == poi)]
    if not df_filtrado.empty:
        media_tempo = df_filtrado['Tempo Permanencia'].mean()
        print(f"\nMédia do Tempo Permanência em {poi} para o dia {data}: {media_tempo:.2f}h")
        return media_tempo
    else:
        print(f"\nNenhum registro encontrado para o dia {data} em {poi}.")
        return 0

def DM(df, local, data):
    df['Data Entrada'] = pd.to_datetime(df['Data Entrada'], format='%d/%m/%Y %H:%M:%S', errors='coerce')
    df['Data Saída'] = pd.to_datetime(df['Data Saída'], format='%d/%m/%Y %H:%M:%S', errors='coerce')
    df_filtrado = df[(df['Data Entrada'].dt.date == data) & (df['Grupo'] == local) & (df['Data Saída'].dt.date == data)]
    if not df_filtrado.empty:
        horas_totais_cm = df_filtrado['Tempo Permanencia'].sum()
        print(f"\nSoma das horas de {local} para o dia {data}: {horas_totais_cm:.2f}h")
        return horas_totais_cm
    else:
        print(f"\nNenhum registro encontrado para o dia {data} em {local}.")
        return 0

def salvar_excel_formatado(data, tpv_ac, dm_rrp):
    caminho_destino = os.path.join(BASE_REPORTS, 'Reports')
    os.makedirs(caminho_destino, exist_ok=True)
    nome_arquivo = 'base de dados reports.xlsx'
    caminho_arquivo = os.path.join(caminho_destino, nome_arquivo)

    nova_linha = pd.DataFrame([{
        'Data': data,
        'TPV AC': tpv_ac,
        'DM RRP': dm_rrp
    }])

    if os.path.exists(caminho_arquivo):
        df_existente = pd.read_excel(caminho_arquivo, engine='openpyxl')
        df_existente['Data'] = pd.to_datetime(df_existente['Data']).dt.date
        nova_data = pd.to_datetime(data).date()
        if nova_data in df_existente['Data'].values:
            print(f"\nA data {nova_data} já existe no arquivo. Nenhuma linha foi adicionada.")
            return caminho_arquivo
        df_resultado = pd.concat([df_existente, nova_linha], ignore_index=True)
    else:
        df_resultado = nova_linha

    with pd.ExcelWriter(caminho_arquivo, engine='openpyxl', mode='w') as writer:
        df_resultado.to_excel(writer, sheet_name='Resumo', index=False)

    print(f"\nRelatório salvo em: {caminho_arquivo}")
    return caminho_arquivo

def contagem_veiculos_por_hora(df, poi, caminho_arquivo_excel):
    df_poi = df[df['Ponto de Interesse'] == poi].copy()
    df_poi['Data Entrada'] = pd.to_datetime(df_poi['Data Entrada'], format='%d/%m/%Y %H:%M:%S', dayfirst=True, errors='coerce')
    df_poi['Data Saída'] = pd.to_datetime(df_poi['Data Saída'], format='%d/%m/%Y %H:%M:%S', dayfirst=True, errors='coerce')

    entradas = df_poi[['Veículo', 'Data Entrada']].copy()
    entradas['Evento'] = 'entrada'
    entradas.rename(columns={'Data Entrada': 'Data Evento'}, inplace=True)

    saidas = df_poi[['Veículo', 'Data Saída']].copy()
    saidas['Evento'] = 'saida'
    saidas.rename(columns={'Data Saída': 'Data Evento'}, inplace=True)

    eventos = pd.concat([entradas, saidas], ignore_index=True)
    eventos.dropna(subset=['Data Evento'], inplace=True)
    eventos.sort_values(by='Data Evento', inplace=True)

    linha_atual = 0
    dentro_poi = set()
    veiculos_no_poi_evento = []

    for _, evento in eventos.iterrows():
        placa = evento['Veículo']
        if evento['Evento'] == 'entrada':
            dentro_poi.add(placa)
            linha_atual += 1
        elif evento['Evento'] == 'saida':
            dentro_poi.discard(placa)
            linha_atual -= 1
        veiculos_no_poi_evento.append(';'.join(sorted(dentro_poi)))

    eventos['Veículos no POI'] = veiculos_no_poi_evento
    df_eventos = eventos.copy()

    start_time = eventos['Data Evento'].min().floor('h')
    end_time = eventos['Data Evento'].max().ceil('h')
    timeline = pd.date_range(start=start_time, end=end_time, freq='h')

    contagem = []
    linha_atual = 0
    dentro_poi = set()
    veiculos_fim_anterior = 0

    for i in range(len(timeline) - 1):
        hora_inicio = timeline[i]
        hora_fim = timeline[i + 1]
        eventos_hora = eventos[(eventos['Data Evento'] >= hora_inicio) & (eventos['Data Evento'] < hora_fim)]
        maximo = minimo = linha_atual

        for _, evento in eventos_hora.iterrows():
            placa = evento['Veículo']
            if evento['Evento'] == 'entrada':
                dentro_poi.add(placa)
                linha_atual += 1
            elif evento['Evento'] == 'saida':
                dentro_poi.discard(placa)
                linha_atual -= 1
            maximo = max(maximo, linha_atual)
            minimo = min(minimo, linha_atual)

        contagem.append({
            'Hora': hora_fim,
            'Veículos no início da hora': veiculos_fim_anterior,
            'Veículos no final da hora': linha_atual,
            'Máximo de veículos': maximo,
            'Mínimo de veículos': minimo,
            'POI': poi,
            'Veículos no POI': ';'.join(sorted(dentro_poi))
        })

        veiculos_fim_anterior = linha_atual

    df_contagem = pd.DataFrame(contagem)

    mes_atual = datetime.now().month
    ano_atual = datetime.now().year

    try:
        wb = load_workbook(caminho_arquivo_excel)
        if 'Candles' in wb.sheetnames:
            df_existente = pd.read_excel(caminho_arquivo_excel, sheet_name='Candles', engine='openpyxl')
            df_existente['Data Evento'] = pd.to_datetime(df_existente['Data Evento'], errors='coerce')
            df_existente = df_existente[~((df_existente['Data Evento'].dt.month == mes_atual) & \
                                         (df_existente['Data Evento'].dt.year == ano_atual) & \
                                         (df_existente['POI'] == poi))]
            df_eventos['POI'] = poi
            df_eventos = pd.concat([df_existente, df_eventos], ignore_index=True)
        else:
            df_eventos['POI'] = poi
    except:
        df_eventos['POI'] = poi

    try:
        if 'Resumo por Hora' in wb.sheetnames:
            df_existente = pd.read_excel(caminho_arquivo_excel, sheet_name='Resumo por Hora', engine='openpyxl')
            df_existente['Hora'] = pd.to_datetime(df_existente['Hora'], errors='coerce')
            df_existente = df_existente[~((df_existente['Hora'].dt.month == mes_atual) & \
                                         (df_existente['Hora'].dt.year == ano_atual) & \
                                         (df_existente['POI'] == poi))]
            df_contagem = pd.concat([df_existente, df_contagem], ignore_index=True)
    except:
        pass

    with pd.ExcelWriter(caminho_arquivo_excel, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_eventos.to_excel(writer, sheet_name='Candles', index=False)
        df_contagem.to_excel(writer, sheet_name='Resumo por Hora', index=False)
        df_eventos.to_excel(writer, sheet_name='Candles', index=False)
        df_contagem.to_excel(writer, sheet_name='Resumo por Hora', index=False)

    print(f"\nContagem por hora salva nas abas 'Candles' e 'Resumo por Hora' em: {caminho_arquivo_excel}")

def aplicar_formatacao_tabela_excel(caminho_arquivo):
    wb = load_workbook(caminho_arquivo)
    for aba in ['Resumo', 'Candles', 'Resumo por Hora']:
        if aba in wb.sheetnames:
            ws = wb[aba]
            max_col = ws.max_column
            max_row = ws.max_row
            if max_row > 1 and max_col > 0:
                col_letter_start = 'A'
                col_letter_end = chr(ord('A') + max_col - 1)
                ref = f"{col_letter_start}1:{col_letter_end}{max_row}"
                nome_tabela = f"Tabela_{aba.replace(' ', '_')}"

                # Verifica se a tabela já existe na planilha
                tabelas_existentes = ws.tables
                if nome_tabela not in tabelas_existentes:
                    table = Table(displayName=nome_tabela, ref=ref)
                    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                    table.tableStyleInfo = style
                    ws.add_table(table)
                else:
                    print(f"Tabela '{nome_tabela}' já existe na aba '{aba}', pulando...")
    wb.save(caminho_arquivo)

def gerar_df_sentinela(poi):
    hoje = datetime.now().date()
    # CORREÇÃO 1: Reduzir janela de tempo para apenas hoje e ontem (evita re-alertas)
    dias = [hoje - timedelta(days=i) for i in range(2)]  # Apenas 2 dias em vez de 4

    df_reports = pd.read_excel(caminho_reports, "Resumo por Hora", engine="openpyxl")
    df_reports["Hora"] = pd.to_datetime(df_reports["Hora"], errors="coerce", dayfirst=True)

    df_filtrado = df_reports[
        (df_reports["POI"] == poi) &
        (df_reports["Hora"].dt.date.isin(dias))
    ].copy()

    return df_filtrado

def identificar_desvios(n_desv, df_desvio):
    df_desvio["n veículos"] = df_desvio["Veículos no POI"].fillna('').apply(
        lambda x: len([v for v in x.split(';') if v])
    )
    df_filtrado = df_desvio[df_desvio["n veículos"] >= n_desv].copy()
    df_filtrado.sort_values("Hora", inplace=True)

    alertas = []
    ultimo_alerta = None
    nivel = 0

    # CORREÇÃO 2: Gerar título baseado apenas na data (sem hora/minuto/segundo)
    def gerar_titulo(poi, hora, nivel):
        data_str = hora.strftime("%d%m%Y")
        # REMOVIDO: hora_str = hora.strftime("%H%M%S")
        return f"RRP_{poi.replace(' ', '')}_N{nivel}_{data_str}"

    for _, row in df_filtrado.iterrows():
        hora_atual = row["Hora"]
        if ultimo_alerta is None or (hora_atual - ultimo_alerta) > timedelta(hours=1):
            nivel = 1
        else:
            nivel = min(nivel + 1, 4)
        ultimo_alerta = hora_atual

        titulo = gerar_titulo(row["POI"], hora_atual, nivel)
        veiculos = str(row["Veículos no POI"]).split(";")

        for v in veiculos:
            alertas.append({
                "Título": titulo,
                "Placa": v.strip(),
                "Ponto_de_Interesse": row["POI"],
                "Data_Hora_Desvio": hora_atual,
                "Data_Hora_Entrada": None,
                "Tempo": None,
                "Nível": f"Tratativa N{nivel}"
            })

    return pd.DataFrame(alertas)

def atualizar_hora_entrada(df_alertas):
    df_candles = pd.read_excel(caminho_reports, sheet_name="Candles", engine="openpyxl")
    df_candles["Data Evento"] = pd.to_datetime(df_candles["Data Evento"], errors="coerce")
    df_entradas = df_candles[df_candles["Evento"].str.lower() == "entrada"]

    def buscar_hora_entrada(row):
        entradas = df_entradas[
            (df_entradas["Veículo"] == row["Placa"]) &
            (df_entradas["POI"] == row["Ponto_de_Interesse"]) &
            (df_entradas["Data Evento"] <= row["Data_Hora_Desvio"])
        ]
        return entradas["Data Evento"].max() if not entradas.empty else pd.NaT

    if df_alertas.empty or df_alertas.columns.size == 0:
        print("DataFrame df_alertas está vazio ou sem colunas.")
        return df_alertas
    else:
        df_alertas["Data_Hora_Entrada"] = df_alertas.apply(buscar_hora_entrada, axis=1)

    def calcular_tempo(row):
        if pd.notnull(row["Data_Hora_Entrada"]):
            delta = datetime.now() - row["Data_Hora_Entrada"]
            return round(delta.total_seconds() / 3600, 2)
        return None

    df_alertas["Tempo"] = df_alertas.apply(calcular_tempo, axis=1)
    return df_alertas

# CORREÇÃO 3: Verificação inteligente no SharePoint
def enviar_para_sharepoint(df):
    ctx = ClientContext(site_url).with_credentials(UserCredential(username, password))
    sp_list = ctx.web.lists.get_by_title(list_name)

    # Obter dados do alerta atual
    hoje_str = datetime.now().strftime("%d%m%Y")
    poi_atual = df.iloc[0]["Ponto_de_Interesse"]
    
    # NOVA VERIFICAÇÃO: Busca alertas existentes hoje para este POI
    existing_items = sp_list.items.select(["Title", "PontodeInteresse"]).top(5000).get().execute_query()
    
    # Verifica se já existe alerta hoje para este POI
    alertas_hoje_poi = [
        item for item in existing_items 
        if hoje_str in item.properties.get("Title", "") 
        and poi_atual in item.properties.get("PontodeInteresse", "")
    ]
    
    if alertas_hoje_poi:
        print(f"Já existe alerta hoje ({hoje_str}) para POI '{poi_atual}'. Nenhum dado foi enviado.")
        return

    # Se não existe, prossegue com o envio
    for _, row in df.iterrows():
        item = {
            "Title": str(row["Título"]),
            "Placa": str(row["Placa"]),
            "PontodeInteresse": str(row["Ponto_de_Interesse"]),
            "Data_Hora_Entrada": row["Data_Hora_Entrada"].strftime('%Y-%m-%dT%H:%M:%S') if pd.notnull(row["Data_Hora_Entrada"]) else None,
            "Tempo": row["Tempo"],
            "Tipo_Alerta": row["Nível"],
            "Status": "Pendente"
        }
        sp_list.add_item(item).execute_query()

    print(f"Desvio enviado para o SharePoint com sucesso. Título: {df.iloc[0]['Título']}")


# ===================================================================
# Função principal: fluxo completo, agora com criação de pasta de mês
# ===================================================================

def main():
    # 1) Define período –– Abril/2025:
    hoje = datetime.today()
    data_inicial = hoje.replace(day=1)
    data_final = hoje
    # data_inicial = datetime.strptime("17/06/2025", "%d/%m/%Y")
    # data_final   = datetime.strptime("17/06/2025", "%d/%m/%Y")

    # 2) Download do relatório C09 (em ~/Downloads/report.xlsx)
    print("Iniciando download do relatório C09...")
    caminho_origem = baixar_relatorio_c09(data_inicial, data_final)

    # 3) Tratamento da planilha e geração do Excel final em BytesIO
    print("Tratando a planilha e adicionando Observações...")
    buffer_resultado = tratar_planilha_c09(caminho_origem)

    # 4) Monta nome do arquivo final (ex.: "C09 01 a 30.04.2025.xlsx")
    data_str = data_final.strftime("%d.%m.%Y")
    nome_arquivo_final = f"C09 01 a {data_str}.xlsx"

    # 5) Monta o nome da pasta no SharePoint para o mês e ano
    MESES_PT = {
        1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
        5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
        9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro"
    }
    numero_mes = data_inicial.month
    nome_pasta_mes = f"{numero_mes:02d}. {MESES_PT[numero_mes]}"  # "01. Janeiro"
    ano_referencia = str(data_inicial.year)

    # 6) Conectar ao SharePoint
    ctx = ClientContext(SITE_URL).with_credentials(UserCredential(SP_USER, SP_PASSWORD))
    caminho_pasta_base = f"{RAIZ_SP}/{BASE_SP}"

    # 7) Verifica/cria a pasta do ano (ex: .../C09/2024)
    full_path_ano = f"{caminho_pasta_base}/{ano_referencia}"
    try:
        pasta_ano = ctx.web.get_folder_by_server_relative_url(full_path_ano)
        ctx.load(pasta_ano)
        ctx.execute_query()
        print(f"Pasta de ano '{ano_referencia}' já existe no SharePoint.")
    except Exception:
        print(f"Pasta de ano '{ano_referencia}' não encontrada. Criando agora…")
        pasta_base = ctx.web.get_folder_by_server_relative_url(caminho_pasta_base)
        pasta_ano = pasta_base.folders.add(ano_referencia)
        ctx.execute_query()
        print(f"Pasta de ano '{ano_referencia}' criada com sucesso.")

    # 8) Verifica/cria a pasta do mês dentro do ano (ex: .../C09/2024/01. Janeiro)
    full_path_mes = f"{full_path_ano}/{nome_pasta_mes}"
    try:
        pasta_mes = ctx.web.get_folder_by_server_relative_url(full_path_mes)
        ctx.load(pasta_mes)
        ctx.execute_query()
        print(f"Pasta '{nome_pasta_mes}' já existe no ano '{ano_referencia}'.")
    except Exception:
        print(f"Pasta '{nome_pasta_mes}' não encontrada em '{ano_referencia}'. Criando agora…")
        pasta_ano = ctx.web.get_folder_by_server_relative_url(full_path_ano)
        pasta_mes = pasta_ano.folders.add(nome_pasta_mes)
        ctx.execute_query()
        print(f"Pasta '{nome_pasta_mes}' criada com sucesso em '{ano_referencia}'.")

    # 9) Exclui versões antigas na pasta correta
    arquivos = pasta_mes.files.get().execute_query()
    for arquivo in arquivos:
        if arquivo.name.startswith("C09 01 a") and arquivo.name.endswith(".xlsx"):
            print("Deletando antigo:", arquivo.name)
            arquivo.delete_object()
    ctx.execute_query()
        
    # 10) Upload do novo arquivo dentro da pasta do mês
    print(f"Enviando '{nome_arquivo_final}' para SharePoint em '{full_path_mes}'…")
    pasta_mes.upload_file(nome_arquivo_final, buffer_resultado.read()).execute_query()
    print("Planilha salva em:", f"{BASE_SP}/{ano_referencia}/{nome_pasta_mes}/{nome_arquivo_final}")

    # 11) Remove o report.xlsx local de ~/Downloads
    try:
        os.remove(caminho_origem)
        print("Arquivo temporário 'report.xlsx' apagado de ~/Downloads.")
    except Exception:
        pass

    print("Teste código Rafael")
    #Inserção código Rafael, para calcular e atualizar a base de dados reports
    caminho_c09_rrp = r'C:\Users\tallespaiva\Suzano S A\Controle operacional - Bases de Dados\CREARE\RRP\C09'
    nome_dia = f"C09 01 a {hoje.strftime('%d.%m.%Y')}.xlsx"
    caminho_final_arquivo_rrp = os.path.join(caminho_c09_rrp, ano_referencia, nome_pasta_mes, nome_dia)
    buffer_resultado.seek(0)
    with open(caminho_final_arquivo_rrp, "wb") as f:
        f.write(buffer_resultado.read()) # ERRO: modo texto
    df = carregar_planilha(caminho_final_arquivo_rrp)
    print(caminho_final_arquivo_rrp)
    print("O código chegou até aqui")
    if df is not None:
            print("Início da atualização de reports")
            ontem = (datetime.now() - timedelta(days=1)).date()
            tpv_ac = TPV(df, "PA AGUA CLARA", ontem)/24
            print("TPV OK")
            totalcmrrp = 91
            dm_valor = DM(df, "Manutenção", ontem)
            print("DM OK ")
            dm_rrp = 100 * (totalcmrrp * 24 - dm_valor - 24*3) / (totalcmrrp * 24)
            caminho_relatorio = salvar_excel_formatado(ontem, tpv_ac, dm_rrp)
            print("Salvar OK")
            contagem_veiculos_por_hora(df, "Descarga Inocencia", caminho_relatorio)
            print("Contagens OK")
            contagem_veiculos_por_hora(df, "Carregamento Fabrica RRP", caminho_relatorio)
            contagem_veiculos_por_hora(df, "PA AGUA CLARA", caminho_relatorio)
            contagem_veiculos_por_hora(df, "Oficina JSL", caminho_relatorio)
            aplicar_formatacao_tabela_excel(caminho_relatorio)
            print("Candles atualizados")

            df_rrp = gerar_df_sentinela("Carregamento Fabrica RRP")
            df_alertas = identificar_desvios(8, df_rrp)
            df_alertas = atualizar_hora_entrada(df_alertas)
            if not df_alertas.empty:
                ultimo_titulo = df_alertas["Título"].iloc[-1]
                df_ultimo = df_alertas[df_alertas["Título"] == ultimo_titulo]
                enviar_para_sharepoint(df_ultimo)

            df_ac = gerar_df_sentinela("PA AGUA CLARA")
            df_alertas = identificar_desvios(8, df_ac)
            df_alertas = atualizar_hora_entrada(df_alertas)
            if not df_alertas.empty:
                ultimo_titulo = df_alertas["Título"].iloc[-1]
                df_ultimo = df_alertas[df_alertas["Título"] == ultimo_titulo]
                enviar_para_sharepoint(df_ultimo)

            df_ino = gerar_df_sentinela("Descarga Inocencia")
            df_alertas = identificar_desvios(15, df_ino)
            df_alertas = atualizar_hora_entrada(df_alertas)
            if not df_alertas.empty:
                ultimo_titulo = df_alertas["Título"].iloc[-1]
                df_ultimo = df_alertas[df_alertas["Título"] == ultimo_titulo]
                enviar_para_sharepoint(df_ultimo)
                
            df_manut = gerar_df_sentinela("Oficina JSL")
            df_alertas = identificar_desvios(15, df_manut)
            df_alertas = atualizar_hora_entrada(df_alertas)
            if not df_alertas.empty:
                ultimo_titulo = df_alertas["Título"].iloc[-1]
                df_ultimo = df_alertas[df_alertas["Título"] == ultimo_titulo]
                enviar_para_sharepoint(df_ultimo)



if __name__ == "__main__":
    main()